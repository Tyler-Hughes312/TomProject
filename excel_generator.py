"""
Excel Generator for Lead Generator
Handles Excel export functionality
"""

import os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from config import EXCEL_COLUMNS, OUTPUT_FOLDER

class ExcelGenerator:
    def __init__(self):
        self.output_folder = OUTPUT_FOLDER
        self.ensure_output_directory()
    
    def ensure_output_directory(self):
        """Ensure output directory exists"""
        os.makedirs(self.output_folder, exist_ok=True)
    
    def export_to_excel(self, businesses: List[Dict], filename: str = None) -> str:
        """Export businesses to Excel file"""
        if not filename:
            filename = f"business_leads_{len(businesses)}_records.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_folder, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Create main sheet
        ws = wb.active
        ws.title = "Business Leads"
        
        # Add headers
        self._add_headers(ws)
        
        # Add data
        self._add_business_data(ws, businesses)
        
        # Create summary sheet
        self._create_summary_sheet(wb, businesses)
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
    
    def _add_headers(self, ws):
        """Add headers to worksheet"""
        headers = EXCEL_COLUMNS
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = {
            'A': 25,  # Business Name
            'B': 35,  # Address
            'C': 15,  # City
            'D': 10,  # State
            'E': 10,  # ZIP Code
            'F': 15,  # Phone Number
            'G': 30,  # Website URL
            'H': 20,  # Business Type
            'I': 10,  # Rating
            'J': 15,  # Review Count
            'K': 10,  # Price Level
            'L': 30   # Yelp URL
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def _add_business_data(self, ws, businesses: List[Dict]):
        """Add business data to worksheet"""
        for row, business in enumerate(businesses, 2):
            ws.cell(row=row, column=1, value=business.get('name', ''))
            ws.cell(row=row, column=2, value=business.get('address', ''))
            ws.cell(row=row, column=3, value=business.get('city', ''))
            ws.cell(row=row, column=4, value=business.get('state', ''))
            ws.cell(row=row, column=5, value=business.get('zip_code', ''))
            ws.cell(row=row, column=6, value=business.get('phone', ''))
            ws.cell(row=row, column=7, value=business.get('website', ''))
            ws.cell(row=row, column=8, value=business.get('business_type', ''))
            ws.cell(row=row, column=9, value=business.get('rating', ''))
            ws.cell(row=row, column=10, value=business.get('review_count', ''))
            ws.cell(row=row, column=11, value=business.get('price_level', ''))
            ws.cell(row=row, column=12, value=business.get('yelp_url', ''))
    
    def _create_summary_sheet(self, wb, businesses: List[Dict]):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary")
        
        # Summary statistics
        total_businesses = len(businesses)
        avg_rating = sum(b.get('rating', 0) for b in businesses) / total_businesses if total_businesses > 0 else 0
        businesses_with_phone = sum(1 for b in businesses if b.get('phone'))
        businesses_with_website = sum(1 for b in businesses if b.get('website'))
        
        # Business type distribution
        business_types = {}
        for business in businesses:
            biz_type = business.get('business_type', 'Unknown')
            business_types[biz_type] = business_types.get(biz_type, 0) + 1
        
        top_business_type = max(business_types.items(), key=lambda x: x[1]) if business_types else ('Unknown', 0)
        
        # City distribution
        cities = {}
        for business in businesses:
            city = business.get('city', 'Unknown')
            cities[city] = cities.get(city, 0) + 1
        
        top_city = max(cities.items(), key=lambda x: x[1]) if cities else ('Unknown', 0)
        
        # Add summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Businesses", total_businesses],
            ["Average Rating", round(avg_rating, 2)],
            ["Businesses with Phone", businesses_with_phone],
            ["Businesses with Website", businesses_with_website],
            ["Top Business Type", f"{top_business_type[0]} ({top_business_type[1]})"],
            ["Top City", f"{top_city[0]} ({top_city[1]})"],
            ["", ""],
            ["Business Type Distribution", ""],
        ]
        
        # Add business type distribution
        for biz_type, count in sorted(business_types.items(), key=lambda x: x[1], reverse=True):
            summary_data.append([biz_type, count])
        
        # Add data to worksheet
        for row, (metric, value) in enumerate(summary_data, 1):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
        
        # Style the summary sheet
        header_font = Font(bold=True)
        for row in range(1, len(summary_data) + 1):
            ws.cell(row=row, column=1).font = header_font
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def export_leads_to_excel(self, leads: List[Dict], filename: str = None) -> str:
        """Export leads to Excel file"""
        if not filename:
            filename = f"leads_export_{len(leads)}_records.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_folder, filename)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Export"
        
        # Add headers
        lead_headers = [
            'Lead ID', 'Business Name', 'Address', 'Phone', 'Website',
            'Status', 'Notes', 'Created Date'
        ]
        
        for col, header in enumerate(lead_headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, lead in enumerate(leads, 2):
            ws.cell(row=row, column=1, value=lead.get('leadid', ''))
            ws.cell(row=row, column=2, value=lead.get('business_name', ''))
            ws.cell(row=row, column=3, value=lead.get('business_address', ''))
            ws.cell(row=row, column=4, value=lead.get('business_phone', ''))
            ws.cell(row=row, column=5, value=lead.get('business_website', ''))
            ws.cell(row=row, column=6, value=lead.get('status', ''))
            ws.cell(row=row, column=7, value=lead.get('notes', ''))
            ws.cell(row=row, column=8, value=lead.get('created', ''))
        
        # Set column widths
        column_widths = [15, 25, 35, 15, 30, 15, 30, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        wb.save(filepath)
        return filepath
