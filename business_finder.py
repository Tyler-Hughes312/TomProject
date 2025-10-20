"""
Business Finder for Lead Generator
Based on existing business finder patterns with optimizations
"""

import os
import asyncio
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass
import requests
from googlemaps import Client
from dotenv import load_dotenv
import json
import logging
# import pandas as pd  # Removed for compatibility
from datetime import datetime
import concurrent.futures
import time
from config import BusinessSearchParams, YELP_API_KEY, GOOGLE_API_KEY
from smarty_verification import get_smarty_verifier

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class VerifiedBusiness:
    """Verified business data structure"""
    name: str
    address: str
    phone_number: str
    city: str
    source: str
    verification_metadata: Dict[str, Any] = None

class BusinessFinder:
    """
    A class to find businesses using Yelp API, cross-reference with Google API,
    optimized for speed without AI verification.
    """
    
    def __init__(self, 
                 yelp_api_key: str = None,
                 google_api_key: str = None):
        """
        Initialize the BusinessFinder with API keys.
        
        Args:
            yelp_api_key: Yelp Fusion API key
            google_api_key: Google Maps API key
        """
        self.yelp_api_key = yelp_api_key or YELP_API_KEY
        self.google_client = Client(google_api_key or GOOGLE_API_KEY)
        
        # API endpoints
        self.yelp_base_url = "https://api.yelp.com/v3"
        
        # Cache for Google API calls
        self.google_cache = {}
        self.google_api_calls = 0
        
        # Performance tracking
        self.start_time = None
        self.api_call_times = []
    
    def search_yelp_businesses(self, params: BusinessSearchParams) -> List[Dict]:
        """
        Search for businesses on Yelp based on given parameters.
        Optimized to minimize API calls by using pagination efficiently.
        
        Args:
            params: BusinessSearchParams object with search criteria
            
        Returns:
            List of business dictionaries from Yelp
        """
        headers = {
            "Authorization": f"Bearer {self.yelp_api_key}",
            "Content-Type": "application/json"
        }
        
        # Convert miles to meters for Yelp API
        radius_meters = int(params.distance_miles * 1609.34)
        
        all_businesses = []
        offset = 0
        limit = min(50, params.max_results)  # Yelp max is 50 per request
        
        while len(all_businesses) < params.max_results:
            url = f"{self.yelp_base_url}/businesses/search"
            params_dict = {
                "term": params.industry,
                "location": params.city,
                "radius": radius_meters,
                "limit": limit,
                "offset": offset,
                "sort_by": "rating"  # Sort by rating
            }
            
            try:
                response = requests.get(url, headers=headers, params=params_dict)
                response.raise_for_status()
                
                data = response.json()
                businesses = data.get("businesses", [])
                
                if not businesses:
                    break  # No more results
                
                all_businesses.extend(businesses)
                offset += len(businesses)
                
                # If we got fewer results than requested, we've reached the end
                if len(businesses) < limit:
                    break
                    
            except requests.exceptions.RequestException as e:
                logger.error(f"Error fetching Yelp data: {e}")
                break
        
        # Limit to requested number of results
        all_businesses = all_businesses[:params.max_results]
        logger.info(f"Found {len(all_businesses)} businesses on Yelp (used {offset//limit + 1} API calls)")
        return all_businesses
    
    def get_google_business_info(self, business_name: str, address: str) -> Optional[Dict]:
        """
        Get business information from Google Places API with caching.
        
        Args:
            business_name: Name of the business
            address: Address of the business
            
        Returns:
            Google business data or None if not found
        """
        # Create cache key
        cache_key = f"{business_name.lower()}_{address.lower()}"
        
        # Check cache first
        if cache_key in self.google_cache:
            return self.google_cache[cache_key]
        
        try:
            # Search for the business
            search_query = f"{business_name} {address}"
            places_result = self.google_client.places(search_query)
            
            self.google_api_calls += 1
            
            if places_result.get("results"):
                place = places_result["results"][0]
                place_id = place.get("place_id")
                
                # Get detailed information
                details_result = self.google_client.place(place_id, fields=[
                    "name", "formatted_address", "formatted_phone_number", 
                    "business_status", "opening_hours", "rating", "user_ratings_total"
                ])
                
                self.google_api_calls += 1
                
                result = details_result.get("result", {})
                
                # Cache the result
                self.google_cache[cache_key] = result
                return result
            
            # Cache None result to avoid repeated failed searches
            self.google_cache[cache_key] = None
            return None
            
        except Exception as e:
            logger.error(f"Error getting Google data for {business_name}: {e}")
            self.google_cache[cache_key] = None
            return None
    
    def fast_verify_business(self, yelp_data: Dict, google_data: Optional[Dict]) -> Dict:
        """
        Fast verification method that compares Yelp and Google data without AI.
        Optimized for speed.
        
        Args:
            yelp_data: Business data from Yelp
            google_data: Business data from Google
            
        Returns:
            Dictionary with verified business information
        """
        # Extract Yelp info
        yelp_info = {
            "name": yelp_data.get("name"),
            "address": yelp_data.get("location", {}).get("address1"),
            "city": yelp_data.get("location", {}).get("city"),
            "phone": yelp_data.get("phone"),
            "rating": yelp_data.get("rating"),
            "review_count": yelp_data.get("review_count")
        }
        
        # Use Yelp data as primary source
        verified_data = {
            "verified_name": yelp_info.get("name"),
            "verified_address": yelp_info.get("address"),
            "verified_city": yelp_info.get("city"),
            "verified_phone": yelp_info.get("phone"),
            "business_status": "unknown",
            "confidence_level": "medium",
            "reasoning": "Using Yelp data as primary source",
            "discrepancies_found": False,
            "discrepancy_details": "Fast verification used"
        }
        
        # Quick Google cross-reference if available
        if google_data:
            # Check for closed business
            if google_data.get("business_status") == "CLOSED_PERMANENTLY":
                verified_data["business_status"] = "closed"
                verified_data["confidence_level"] = "high"
                verified_data["reasoning"] = "Google indicates business is permanently closed"
            
            # Use Google phone if Yelp doesn't have one
            if not verified_data["verified_phone"] and google_data.get("formatted_phone_number"):
                verified_data["verified_phone"] = google_data["formatted_phone_number"]
        
        return verified_data
    
    def process_business_batch(self, yelp_businesses: List[Dict]) -> List[VerifiedBusiness]:
        """
        Process a batch of businesses in parallel for maximum speed.
        
        Args:
            yelp_businesses: List of Yelp business data
            
        Returns:
            List of verified business objects
        """
        verified_businesses = []
        
        # Process businesses in parallel using ThreadPoolExecutor
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            # Submit all tasks
            future_to_business = {}
            for yelp_business in yelp_businesses:
                future = executor.submit(self._process_single_business, yelp_business)
                future_to_business[future] = yelp_business
            
            # Collect results as they complete
            for future in concurrent.futures.as_completed(future_to_business):
                try:
                    business = future.result()
                    if business:
                        verified_businesses.append(business)
                except Exception as e:
                    yelp_business = future_to_business[future]
                    logger.error(f"Error processing business {yelp_business.get('name', 'Unknown')}: {e}")
        
        return verified_businesses
    
    def _process_single_business(self, yelp_business: Dict) -> Optional[VerifiedBusiness]:
        """
        Process a single business (used by parallel processing).
        
        Args:
            yelp_business: Yelp business data
            
        Returns:
            VerifiedBusiness object or None if processing failed
        """
        try:
            # Get Google data for cross-referencing
            yelp_address = yelp_business.get("location", {}).get("address1", "")
            google_data = self.get_google_business_info(
                yelp_business.get("name", ""),
                yelp_address
            )
            
            # Fast verification
            verified_data = self.fast_verify_business(yelp_business, google_data)
            
            # Create verification metadata
            verification_metadata = {
                "confidence_level": verified_data.get("confidence_level"),
                "business_status": verified_data.get("business_status"),
                "discrepancies_found": verified_data.get("discrepancies_found"),
                "discrepancy_details": verified_data.get("discrepancy_details"),
                "yelp_rating": yelp_business.get("rating"),
                "yelp_review_count": yelp_business.get("review_count")
            }
            
            # Create VerifiedBusiness object
            business = VerifiedBusiness(
                name=verified_data.get("verified_name"),
                address=verified_data.get("verified_address"),
                phone_number=verified_data.get("verified_phone"),
                city=verified_data.get("verified_city"),
                source=f"Yelp+Google_Fast_Verified_{verified_data.get('confidence_level', 'unknown')}",
                verification_metadata=verification_metadata
            )
            
            return business
            
        except Exception as e:
            logger.error(f"Error processing business {yelp_business.get('name', 'Unknown')}: {e}")
            return None
    
    def find_and_export_businesses(self, params: BusinessSearchParams, 
                                 filename: Optional[str] = None) -> str:
        """
        Main method to find businesses and export to Excel immediately.
        Optimized for speed with parallel processing.
        
        Args:
            params: BusinessSearchParams object with search criteria
            filename: Optional custom filename
            
        Returns:
            Path to the created Excel file
        """
        self.start_time = time.time()
        logger.info(f"Starting fast business search for {params.industry} in {params.city}")
        
        # Reset counters
        self.google_api_calls = 0
        self.google_cache.clear()
        
        # Step 1: Get businesses from Yelp
        yelp_businesses = self.search_yelp_businesses(params)
        
        if not yelp_businesses:
            logger.warning("No businesses found on Yelp")
            return ""
        
        # Step 2: Process businesses in parallel
        logger.info(f"Processing {len(yelp_businesses)} businesses in parallel...")
        verified_businesses = self.process_business_batch(yelp_businesses)
        
        # Step 3: Export to Excel immediately
        excel_file = self.export_to_excel(verified_businesses, params, filename)
        
        # Log performance metrics
        total_time = time.time() - self.start_time
        logger.info(f"Completed in {total_time:.2f} seconds")
        logger.info(f"Processed {len(verified_businesses)} businesses")
        logger.info(f"Total Google API calls: {self.google_api_calls}")
        logger.info(f"Average time per business: {total_time/len(verified_businesses):.2f} seconds")
        
        return excel_file
    
    def export_to_excel(self, businesses: List[VerifiedBusiness], params: BusinessSearchParams, 
                       filename: Optional[str] = None) -> str:
        """
        Export business results to an Excel file with comprehensive information.
        
        Args:
            businesses: List of verified business objects
            params: Search parameters used
            filename: Optional custom filename (will generate one if not provided)
            
        Returns:
            Path to the created Excel file
        """
        if not businesses:
            logger.warning("No businesses to export")
            return ""
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_city = params.city.replace(" ", "_").replace(",", "")
            safe_industry = params.industry.replace(" ", "_")
            filename = f"businesses_{safe_industry}_{safe_city}_{timestamp}.xlsx"
        
        # Prepare data for Excel
        excel_data = []
        
        for business in businesses:
            metadata = business.verification_metadata or {}
            
            row = {
                "Business Name": business.name,
                "Address": business.address,
                "City": business.city,
                "Phone Number": business.phone_number,
                "Source": business.source,
                "Confidence Level": metadata.get("confidence_level", "unknown"),
                "Business Status": metadata.get("business_status", "unknown"),
                "Yelp Rating": metadata.get("yelp_rating", "N/A"),
                "Yelp Review Count": metadata.get("yelp_review_count", "N/A"),
                "Discrepancies Found": metadata.get("discrepancies_found", False),
                "Discrepancy Details": metadata.get("discrepancy_details", ""),
                "Verification Reasoning": metadata.get("reasoning", ""),
            }
            excel_data.append(row)
        
        # Create Excel file using openpyxl directly
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Business Data"
        
        # Add headers
        headers = list(excel_data[0].keys()) if excel_data else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add data
        for row, business in enumerate(excel_data, 2):
            for col, (key, value) in enumerate(business.items(), 1):
                ws.cell(row=row, column=col, value=value)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Metric")
        ws_summary.cell(row=1, column=2, value="Value")
        
        summary_data = [
            ("Total Businesses Found", len(businesses)),
            ("Search City", params.city),
            ("Industry/Category", params.industry),
            ("Search Radius (miles)", params.distance_miles),
            ("Max Results Requested", params.max_results),
            ("Search Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Google API Calls", self.google_api_calls),
            ("Google Cache Hits", self.get_api_usage_stats()["google_cache_hits"]),
            ("Google Cache Misses", self.get_api_usage_stats()["google_cache_misses"])
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1, value=metric)
            ws_summary.cell(row=row, column=2, value=value)
        
        # Save the workbook
        wb.save(filename)
        
        logger.info(f"Excel file created: {filename}")
        logger.info(f"Exported {len(businesses)} businesses to {filename}")
        
        return filename
    
    def get_api_usage_stats(self) -> Dict[str, Any]:
        """
        Get statistics about API usage for optimization tracking.
        
        Returns:
            Dictionary with API usage statistics
        """
        return {
            "google_api_calls": self.google_api_calls,
            "google_cache_size": len(self.google_cache),
            "google_cache_hits": len([v for v in self.google_cache.values() if v is not None]),
            "google_cache_misses": len([v for v in self.google_cache.values() if v is None])
        }
    
    def clear_google_cache(self):
        """Clear the Google API cache to free memory."""
        self.google_cache.clear()
        logger.info("Google API cache cleared")
    
    @classmethod
    def from_env(cls):
        """
        Create a BusinessFinder instance using environment variables.
        
        Returns:
            BusinessFinder instance
            
        Raises:
            ValueError: If required environment variables are missing
        """
        yelp_api_key = os.getenv('YELP_API_KEY')
        google_api_key = os.getenv('GOOGLE_API_KEY')
        
        if not yelp_api_key:
            raise ValueError("YELP_API_KEY environment variable is required")
        if not google_api_key:
            raise ValueError("GOOGLE_API_KEY environment variable is required")
        
        return cls(
            yelp_api_key=yelp_api_key,
            google_api_key=google_api_key
        )
    
    def verify_business_address(self, business: Dict) -> Dict:
        """
        Verify business address using Smarty Streets
        
        Args:
            business: Business dictionary with address information
            
        Returns:
            Business dictionary with verification data added
        """
        verifier = get_smarty_verifier()
        if not verifier:
            logger.warning("Smarty Streets verifier not available")
            return business
        
        location = business.get('location', {})
        street = location.get('address1', '')
        city = location.get('city', '')
        state = location.get('state', '')
        zip_code = location.get('zip_code', '')
        
        if not all([street, city, state, zip_code]):
            logger.warning(f"Incomplete address for business {business.get('name', 'Unknown')}")
            return business
        
        # Verify the address
        verification_result = verifier.verify_address(street, city, state, zip_code)
        
        # Add verification data to business
        business['address_verified'] = verification_result['verified']
        business['address_verification_status'] = verification_result['status']
        business['verified_address'] = verification_result['verified_address']
        business['verified_city'] = verification_result['verified_city']
        business['verified_state'] = verification_result['verified_state']
        business['verified_zip_code'] = verification_result['verified_zip_code']
        business['verification_confidence'] = verification_result['confidence']
        
        if verification_result['error']:
            business['verification_error'] = verification_result['error']
        
        return business
